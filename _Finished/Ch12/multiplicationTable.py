"""Create a program that takes a number N from teh command line and creates an NxN multiplication table in an Excel spreadsheet.
    ex.
    [  ][ 2][ 3][ 4]
    [ 2][ 4][ 6][ 8]
    [ 3][ 6][ 9][12]
    [ 4][ 8][12][16]
"""
import openpyxl
from openpyxl.styles import Font
import os

def create_multiplication_table(max_input):
    if type(max_input) != int:
        print("Not an int!")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    bold_font = Font(bold=True)
    max_input += 1

    # A1, B1, C1, etc
    for i in range(1, max_input):
        current_letter = ws.cell(row=1, column=i + 1).column_letter
        ws[current_letter + "1"] = i
        ws[current_letter + "1"].font = bold_font

    # A1, A2, A3, etc.
    for i in range(1, max_input):
        ws["A" + str(i + 1)] = i
        ws["A" + str(i + 1)].font = bold_font

    max_input += 1
    # Everything inside the ^ boundaries
    current_col = 2
    for row_loop in range(2, max_input):
        for col_loop in range(2, max_input):
            current_letter = ws.cell(row=row_loop, column=col_loop).column_letter
            # Output: =(A[X]*[Y]1) ; x = row, y = col
            ws[current_letter + str(row_loop)] = f"=(A{str(row_loop)}*{current_letter}1)"
    

    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    wb.save("Workfiles/MultiplicationTable.xlsx")


if __name__ == "__main__":
    try:
        while True:
            user_input = int(input("Enter an integer: "))
            if user_input < 9000 and user_input >= 1:
                break
            else:
                print("Int is too high, try again.")

    except ValueError:
        user_input = "Oh god it's gonna crash."
    # user_input = 1000
    create_multiplication_table(user_input)
