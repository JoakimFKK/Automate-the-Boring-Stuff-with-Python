import os
import openpyxl

os.chdir(os.path.dirname(os.path.realpath(__file__)))
print(os.getcwd())

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>


"""Get sheets from the workbook"""
sheet = wb['Sheet1']  # Bogen bruger en deprecated function.
print(sheet['A1'])  # <Cell 'Sheet1'.A1>
print(sheet['A1'].value)  # 2015-04-05 13:34:02

c = sheet['B1']
print(c.value)  # Apples
print(sheet['C1'].value)  # 73

"""Get cells by cells and rows"""
print(sheet.cell(row=1, column=2)) # <Cell 'Sheet1'.B1>
print(sheet.cell(row=1, column=2).value) # Apples

for i in range(1, 8, 2):
    """Output:
        1 Apples
        3 Pears
        5 Apples
        7 Strawberries
    """
    print(i, sheet.cell(row=i, column=2).value)

"""You can determine the size of the sheet with the attributes:"""
print(sheet.max_row) # 7
print(sheet.max_column) # 3
