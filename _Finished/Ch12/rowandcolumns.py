import openpyxl

def setting_row_and_column_width_height():
    """row_dimensions['[COL/ROW THAT IS AFFECTED]']"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = "Tall row"
    sheet['B2'] = "Wide column"

    sheet.row_dimensions[1].height = 70 
    sheet.column_dimensions['B'].width = 20 
    wb.save("Ch12/Workfiles/rowandcolumnsDOTpy.xlsx")

def merging_and_unmerging_cells():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.merge_cells('A1:D3')
    sheet['A1'] = "Big boy coming through"
    sheet.merge_cells('C5:D5')
    sheet['C5'] = "Smaller boy also here"
    wb.save("Ch12/Workfiles/rowandcolumnsDOTpy.xlsx")

def freezing_panes():
    wb = openpyxl.load_workbook("Ch12/Workfiles/NEW_PrduceSale.xlsx")
    sheet = wb.active
    sheet.freeze_panes = 'A2'
    wb.save("Ch12/Workfiles/FROZEN_PrduceSale.xlsx")


freezing_panes()
# setting_row_and_column_width_height()
# merging_and_unmerging_cells()