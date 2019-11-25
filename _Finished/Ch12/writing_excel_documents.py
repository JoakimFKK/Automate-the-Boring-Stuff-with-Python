import openpyxl

"""1ST BIT"""
# wb = openpyxl.Workbook()
# print(wb.sheetnames) # Output: ['Sheet']
# sheet = wb.active
# print(sheet.title) # Sheet
# sheet.title = 'ACAB'
# print(wb.sheetnames) # ['ACAB']

"""Creating and saving Excel documents."""
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "ACAB"
# wb.save("acab.xlsx") # Output: An empty excel file, med sheet title "ACAB"

"""Creating and removing sheets."""
# wb = openpyxl.Workbook()
# print(wb.sheetnames)  # ['Sheet']
# wb.create_sheet()
# print(wb.sheetnames)  # ['Sheet', 'Sheet1']
# wb.create_sheet(index=0, title="First sheet")
# print(wb.sheetnames)  # ['First sheet', 'Sheet', 'Sheet1']
# wb.create_sheet(index=2, title="Middle sheet")
# print(wb.sheetnames)  # ['First sheet', 'Sheet', 'Middle sheet', 'Sheet1']
# del wb['Sheet1']
# print(wb.sheetnames)  # ['First sheet', 'Sheet', 'Middle sheet']

"""Writing Values to cells"""
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = "Hello world"
print(sheet['A1'].value) # Hello world
print(sheet['A2'].value) # None