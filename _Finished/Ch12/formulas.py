import openpyxl

def create_formula():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 200
    sheet['A2'].value = 300
    sheet['A3'].value = '=SUM(A1:A2)'
    wb.save("Ch12/Workfiles/formulasDOTpy formula.xlsx")

def get_literal_formula():
    wb = openpyxl.load_workbook("Ch12/Workfiles/formulasDOTpy formula.xlsx")
    sheet = wb.active
    return sheet['A3'].value #  '=SUM(A1:A2)'

def get_data_formula():
    wb = openpyxl.load_workbook("Ch12/Workfiles/formulasDOTpy formula.xlsx", data_only=True)
    sheet = wb.active
    print(sheet['A1'].value) #  200
    print(sheet['A2'].value) #  300
    print(sheet['A3'].value) #  None
    return sheet['A3'].value

print("Getting the formula: \t", get_literal_formula())
print("Getting the data: \t", get_data_formula())